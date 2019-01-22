#Colin Kennelly
import openpyxl


def census_data():
    cen_data = openpyxl.load_workbook("massachusetts_population_1980-2010.xlsx")
    cen_sheet = cen_data.get_active_sheet()
    mass_data = openpyxl.load_workbook("MAEmplyomentData.xlsx")
    mass_sheet = mass_data.get_active_sheet()
    for marow in mass_sheet.iter_rows(min_row=2):
        for cenrow in cen_sheet.iter_rows(min_row=12):
            ma_city = ""
            if marow[0].value:
                ma_city = marow[0].value
            cen_city = ""
            if cenrow[3].value:
                cen_city = cenrow[3].value
            if ma_city.strip() == cen_city.strip():
                nolabor = cenrow[8].value - marow[1].value
                print(f"{ma_city} has {cenrow[8].value} citizens, with {marow[1].value} "
                      f"laborers, and {nolabor} citizens currently not in the work force")
census_data()

